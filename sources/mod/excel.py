'''
Corresponde a las funciones que tienen relación con hojas de cálculo. Funciones que buscan datos, lista de datos, crean y eliminan archivos, hojas, etc...
'''
#from datetime import datetime, date, time
#import calendar
#from datetime import timedelta
# Abre, modifica y guarda los cambios en archivos excel
import openpyxl as op
# Importado para convertir csv a xlsx
import pandas as pd

# Devuelve 7 listas que tienen lo siguiente:
    # Lista1 = Línea
    # Lista2 = Codigo
    # Lista3 = Descripción
    # Lista4 = PSPV
    # Lista5 = Precio Base (POR EL MOMENTO MONOTRIBUTO INDICADO POR PARÁMETRO)
    # Lista6 = Puntos
    # Lista7 = Puntos Me Gusta

# La función extraerá la cantidad de datos posible, y cuando se encuentre con 10 celdas vacías consecutivas, terminará. Aún así sólo cargará los datos de aquellas filas que
    # tengan algún código cargado.
# IMPORTANTE: Para el módulo, la fila 1 del archivo = 0.
def Dev_Listas(Nom_Libro, Nom_Hoja, ColumnaPrecio):
    Aviso = 0
    try:
        wb = op.load_workbook(Nom_Libro, data_only=True)
        Aviso += 1
        Hoja = wb.get_sheet_by_name(Nom_Hoja)
        Aviso += 1

        Col_1 = Hoja['B']
        Col_2 = Hoja['C']
        Col_3 = Hoja['D']
        Col_4 = Hoja['F']
        Col_5 = Hoja[ColumnaPrecio]
        Col_6 = Hoja['E']
        Col_7 = Hoja['K']

        Lista1 = []
        Lista2 = []
        Lista3 = []
        Lista4 = []
        Lista5 = []
        Lista6 = []
        Lista7 = []
        Lista8 = []

        Tope = len(Col_2)
        Cont = 0
        while Cont < Tope:
            Valor2 = Col_2[Cont].value
            if Valor2 != None:
                if Es_Numero_Int(Valor2) == True:
                    # Columna de Linea
                    Lista1.append(Col_1[Cont].value)
                    # Columna de Codigo
                    Lista2.append(str(Valor2))
                    # Columna de Descripcion
                    Lista3.append(Col_3[Cont].value)
                    # Columna de PSPV
                    if Col_4[Cont].value != None:
                        Lista4.append(float(Col_4[Cont].value))
                    else:
                        Lista4.append(round((float(Col_5[Cont].value)) / 0.73))
                    # Columna de Monotributo
                    Lista5.append(float(Col_5[Cont].value))
                    # Columna de Puntos
                    if Col_6[Cont].value != None:
                        Lista6.append(int(Col_6[Cont].value))
                    else:
                        Lista6.append(0)
                    # Columna de Puntos MG
                    Lista7.append(0)
                    Lista8.append(Cont+1)
            Cont += 1
            #print(Cont)
        return Lista1, Lista2, Lista3, Lista4, Lista5, Lista6, Lista7, Lista8
    except:
        return Aviso

# La función extraerá la cantidad de datos posible, y cuando se encuentre con 10 celdas vacías consecutivas, terminará. Aún así sólo cargará los datos de aquellas filas que
    # tengan algún código cargado.
# IMPORTANTE: Para el módulo, la fila 1 del archivo = 0.
def Dev_Listas_Contactos(Nom_Libro, Nom_Hoja):
    Aviso = 0
    try:
        wb = op.load_workbook(Nom_Libro, data_only=True)
        Aviso += 1
        Hoja = wb.get_sheet_by_name(Nom_Hoja)
        Aviso += 1

        Col_0 = Hoja['A']   # Modo agendado
        Col_1 = Hoja['B']   # Nombre1
        Col_2 = Hoja['C']   # Nombre2
        Col_3 = Hoja['D']   # Apellido
        Col_4 = Hoja['E']   # E, F, G, H, I, J: Nombre3
        Col_5 = Hoja['F']   
        Col_6 = Hoja['G']
        Col_7 = Hoja['H']
        Col_8 = Hoja['I']
        Col_9 = Hoja['J']
        Col_10 = Hoja['AG'] # Celular
        Col_11 = Hoja['AI'] # Dirección
        Col_12 = Hoja['AR'] # Cuenta Bussines

        Lista1 = []
        Lista2 = []
        Lista3 = []
        Lista4 = []
        Lista5 = []
        Lista6 = []
        Lista7 = []
        Lista8 = []

        Tope = len(Col_0)
        Cont = 1
        while Cont < Tope:
            # Columna de Modo Agendado
            if Col_0[Cont].value != None:
                Lista1.append(Col_0[Cont].value)
            else:
                Lista1.append("")
            # Columna de Nombre1
            if Col_1[Cont].value != None:
                Lista2.append(Col_1[Cont].value)
            else:
                Lista2.append("")
            # Columna de Nombre2
            if Col_2[Cont].value != None:
                Lista3.append(Col_2[Cont].value)
            else:
                Lista3.append("")
            # Columna de Apellido
            if Col_3[Cont].value != None:
                Lista4.append(Col_3[Cont].value)
            else:
                Lista4.append("")
            # Columna de Nombre3
            aux = ""
            if Col_4[Cont].value != None:
                aux = Col_4[Cont].value
            if Col_5[Cont].value != None:
                aux += " " + Col_5[Cont].value
            if Col_6[Cont].value != None:
                aux += " " + Col_6[Cont].value
            if Col_7[Cont].value != None:
                aux += " " + Col_7[Cont].value
            if Col_8[Cont].value != None:
                aux += " " + Col_8[Cont].value
            if Col_9[Cont].value != None:
                aux += " " + Col_9[Cont].value
            Lista5.append(aux)
            # Columna de Celular
            if Col_10[Cont].value != None:
                Lista6.append(Col_10[Cont].value)
            else:
                Lista6.append("")
            # Columna de Dirección
            if Col_11[Cont].value != None:
                Lista7.append(Col_11[Cont].value)
            else:
                Lista7.append("")
            # Columna de Cuenta Bussines
            if Col_12[Cont].value != None:
                Lista8.append(Col_12[Cont].value)
            else:
                Lista8.append("")
            Cont += 1
            #print(Cont)
        return Lista1, Lista4, Lista2, Lista3, Lista5, Lista6, Lista7, Lista8
    except:
        return Aviso

# Determina si es un número int.
# Devuelve: V o F
def Es_Numero_Int(Valor):
    try:
        resultado = int(Valor)
        return True
    except ValueError:
        return False

# Devuelve una lista, que contiene todos los datos de una columna dada. Comienza desde la fila que se indica por parámetro y termina donde se indique por parámetro.
# Si no se indica el final de la fila, por defecto se utiliza la última celda con valor.
# IMPORTANTE: Los parámetros Fila_Ini y Fila_Fin, corresponden a los valores reales dentro de la planilla. Se aclara ésto porque para el módulo, la fila 1 del archivo = 0.
def Dev_Listas2(Nom_Libro, Nom_Hoja, Nom_Col = 'A', Fila_Ini = 2, Fila_Fin = -1):
    Aviso = -3
    try:
        # Cargamos en la variable PrimerCol toda la columna con valores que se indica. Esta función carga toda la columna aunque no tenga datos, si por ejemplo el que creó la
            # hoja de excel puso un formato en las primeras 1000 filas, ésto va a interpretar que la columna mide 1000 filas, aunque no tengan datos.
        wb = op.load_workbook(Nom_Libro)
        Aviso += 1
        Hoja = wb.get_sheet_by_name(Nom_Hoja)
        Aviso += 1
        PrimerCol = Hoja[Nom_Col]
        Largo = len(PrimerCol)
        Vacio = False
        Tope = 0

        if Fila_Fin == -1:
            Vacio = True
            Tope = Largo
        else:
            Tope = Fila_Fin
        Cont = Fila_Ini - 1

        Lista = []

        while Cont < Tope:
            if PrimerCol[Cont].value != None:
                Lista.append(PrimerCol[Cont].value)
                Cont += 1
            else:
                if Vacio == True:
                    return Lista
                else:
                    Lista.append('')
        return Lista
    except:
        return Aviso


# Escribe un dato en una celda determinada
# Devuelve True or False si se pudo o no realizar la operación. Habitualmente no se puede cuando el libro está abierto por otro usuario.
def Escribe_Dato(Valor, Nom_Libro, Nom_Hoja, Nom_Col, Num_Fila):
    try:
        wb = op.load_workbook(Nom_Libro)
        Hoja = wb.get_sheet_by_name(Nom_Hoja)
        Celda = str(Nom_Col + str(Num_Fila))
        Hoja[Celda] = int(Valor)
        wb.save(Nom_Libro)
        return True
    except:
        return False

# Se usa para escribir una serie de datos, un cuadro entero o bien una cantidad de columnas indefinidas, desde 1 hasta n elementos, y lo mismo para las filas. Debe venir 
    # una lista con listas que representan todos los datos a cargar. Si suponemos que queremos copiar una matriz de datos de 4 columnas por 3 filas desde B2 hasta E4, 
    # entonces los argumentos serían así:
    # Nom_Libro: >>> 'Seguimiento 2020.xlsx' .Es el nombre del libro que se desea escribir.
    # Nom_Hoja: >>> 'Auxiliar' .Es el nombre de la hoja donde se desea escribir.
    # Matriz: >>> Lista_principal[Lista1, Lista2, Lista3, Lista4]. Cada lista corresponde a una columna, por ende, cada lista debe contener 3 datos.
    # Col_Ini: >>> 2 .Debe tener el valor de la columna ya sea con número o letras.
    # Fila_Ini: >>> 2 .Es la fila de inicio representada por un valor Entero.
        # NOTA: Tener en cuenta que las coordenadas de Columna y Fila corresponden a la esquina superior izquierda de la matriz a copiar.
    # IMPORTANTE:
        # La posición 0 (cero) de cada lista interna (en ésta caso son las listas 'Lista1', 'Lista2', etc), indica qué se quiere colocar en los espacios vacíos de su columna.
        # Es decir, que si por ejemplo la primer columna corresponde a una serie de datos de Nombres de productos, personas, etc, se puede colocar un vacío que sería ''.
        # Pero tal vez la segunda columna contiene valores enteros, y deseamos que haya un cero, entonces colocamos 0, pero en formato de Entero, no confundir con '0' de 
            # formato string.
        # Adicionalmente, se puede poner un valor para que se rellenen los espacios vacíos en él, como por ejemplo si tenemos una columna de fechas y colocamos en la posición
            # [0] de la lista una fecha 03-03-2020, entonces asigna dicho valor a cada espacio vacío.
# Devuelve 3 valores que representan lo siguiente:
    # 0: >>> No se pudo realizar la operación, lo más probable que no se haya podido abrir el archivo.
    # 1: >>> Guardado con éxito.
    # 2: >>> Se han cargado cambios pero no se han podido guardar. Lo más probable es que el archivo esté abierto por algún usuario.
def Escribe_Listas( Nom_Libro, Nom_Hoja, Matriz, Col_Ini, Fila_Ini):
    Iniciado = False
    try:
        # Se abre el archivo excel y se cargan las variables con el archivo y la hoja que se va a escribir
        wb = op.load_workbook(Nom_Libro)
        Hoja = wb.get_sheet_by_name(Nom_Hoja)

        # Damos valores a las variables que nos van a ubicar tanto en la matriz a copiar como en las celdas a escribir.
        Cant_Col = len(Matriz)
        Cant_Fila = len(Matriz[0])
        Contador_Col = 0
        Fila_Ini -= 1

        # Bucle que recorre tanto la matriz como las celdas a copiar, ya que obviamente son iguales en tamaño.
        # IMPORTANTE: Para analizar el bucle, hay que recordar que la primer posición de las listas es un valor de configuración, y no es un valor para cargar, por eso hay una
            # fila menos a tener en cuenta.
        while Contador_Col < Cant_Col:
            Contador_Fila = 1
            while Contador_Fila < Cant_Fila:
                Valor = 0
                if Matriz[Contador_Col][Contador_Fila] == '':
                    Valor = Matriz[Contador_Col][0]
                else:
                    Valor = Matriz[Contador_Col][Contador_Fila]
                Hoja.cell(row = Fila_Ini + Contador_Fila, column = Col_Ini + Contador_Col).value = Valor
                Iniciado = True
                Contador_Fila += 1
            Contador_Col += 1

        # Guardamos los cambios, este momento es clave porque todo lo anterior se puede ejecutar siempre que el archivo exista, sin importar si está o no abierto por otros
            # usuarios, pero, los cambios no se pueden guardar si el archivo está abierto así que ésta acción finaliza con la operación y define si fue o no exitosa.
        wb.save(Nom_Libro)
        return 1
    except:
        if Iniciado == True:
            return 2
        else:
            return 0


# Función que copia una lista donde se indique, pero si en el lugar existe una lista con uno o más parámetros que coincidan, entonces sólo suma los valores
# Devuelve 3 valores que representan lo siguiente:
    # 0: >>> No se pudo realizar la operación, lo más probable que no se haya podido abrir el archivo.
    # 1: >>> Guardado con éxito.
    # 2: >>> Se han cargado cambios pero no se han podido guardar. Lo más probable es que el archivo esté abierto por algún usuario.
def Escribe_Lista_Suma_Coincidencias(Nom_Libro, Nom_Hoja, Matriz, Col_Ini, Fila_Ini):
    Iniciado = False
    try:
        # Se abre el archivo excel y se cargan las variables con el archivo y la hoja que se va a escribir
        wb = op.load_workbook(Nom_Libro)
        Hoja = wb.get_sheet_by_name(Nom_Hoja)
        Iniciado = True

        # Damos valores a las variables que nos van a ubicar tanto en la matriz a copiar como en las celdas a escribir.
        Cant_Col = len(Matriz)
        Cant_Fila = len(Matriz[0])

        Contador_Col = 0
        Contador_Fila = 1

        # Recorremos la primer columna para determinar con la variable Primer_Fila_Vacia, si es que el archivo está limpio (Cuando Primer_Fila_Vacia == 0), pero al 
            # mismo tiempo necesitaremos saber cuál es la primer fila vacía del excel para cargar los nuevos clientes de la fecha (Cuando Primer_Fila_Vacia > Fila_Ini).
            # Adicionalmente capturamos el valor de la última fecha cargada necesaria más adelante.
        Bucle = True
        Primer_Fila_Vacia = Fila_Ini
        Ultima_Fecha = ''
        while Bucle == True:
            ValorHoja = Hoja.cell(row = Primer_Fila_Vacia, column = Col_Ini).value
            if ValorHoja == None:
                if Primer_Fila_Vacia == Fila_Ini:
                    Ultima_Fecha = 0
                else:
                    Ultima_Fecha = Hoja.cell(row = Primer_Fila_Vacia - 1, column = Col_Ini).value
                Bucle = False
                break
            Primer_Fila_Vacia += 1
        
        # Bucle para recorrer FILA x FILA la matriz de datos, comparar nombres y fechas para actualizar datos, o cargar nuevos clientes
        # Consta de 2 partes, en la primera le indicamos si tiene que sumar datos a algo antiguo y su fila, y en la segunda parte directamente cargamos los datos
        while Contador_Fila < Cant_Fila:
      
            
            Sobreescribe = False
            Fila = 0

            # Una vez ingresado a la fila, controlamos que ese dato no corresponda con anterioridad a otro cliente, por eso vamos a comparar el cliente nuevo en la lista
                # de clientes, y si al mismo tiempo comparte con la fecha de hoy.
                # Debido a que estos datos en la mayoría de los casos son necesarios mas adelante, recolectamos todos los datos de la fila en éste momento.
            Fecha_Nueva = Matriz[0][Contador_Fila]
            Cliente_Nuevo = Matriz[1][Contador_Fila]
            Drop_Nuevo = Matriz[2][Contador_Fila]
            Pagos_Nuevo = Matriz[3][Contador_Fila]
            Nuevo_VoF = Matriz[4][Contador_Fila]
            
            # Tener en cuenta que no llegan nombres repetidos, si el Cajero cargó 2 veces importes a un cliente, se ajustaron durante la creación de la lista
            Fila = Primer_Fila_Vacia
            if Fecha_Nueva == Ultima_Fecha:
                # Bucle para buscar si ya tuvimos una carga anterior del cliente
                Inicio = 2
                while Inicio < Primer_Fila_Vacia:
                    if Hoja.cell(row = Inicio, column = 2).value == Cliente_Nuevo:
                        Sobreescribe = True
                        Fila = Inicio
                    Inicio += 1

            # 2). Cargamos los datos
            # True: Sobreescribimos los datos del cliente repetido. False: Creamos un registro nuevo
            if Sobreescribe == True:
                if Drop_Nuevo != '':
                    if Drop_Nuevo > 0:
                        Hoja.cell(row = Fila, column = 3).value = Hoja.cell(row = Fila, column = 3).value + Drop_Nuevo
                if Pagos_Nuevo != '':
                    if Pagos_Nuevo > 0:
                        Hoja.cell(row = Fila, column = 4).value = Hoja.cell(row = Fila, column = 4).value + Pagos_Nuevo
            else:
                Hoja.cell(row = Fila, column = 1).value = Fecha_Nueva
                Hoja.cell(row = Fila, column = 2).value = Cliente_Nuevo
                if Drop_Nuevo != '':
                    Hoja.cell(row = Fila, column = 3).value = Drop_Nuevo
                else:
                    Hoja.cell(row = Fila, column = 3).value = Matriz[2][0]
                if Pagos_Nuevo != '':
                    Hoja.cell(row = Fila, column = 4).value = Pagos_Nuevo
                else:
                    Hoja.cell(row = Fila, column = 4).value = Matriz[3][0] 
                if Nuevo_VoF == True:
                    Hoja.cell(row = Fila, column = 5).value = 'CLIENTE NUEVO CREADO POR EL CAJERO'
                Primer_Fila_Vacia += 1
            Contador_Fila += 1




        # Guardamos los cambios, este momento es clave porque todo lo anterior se puede ejecutar siempre que el archivo exista, sin importar si está o no abierto por otros
            # usuarios, pero, los cambios no se pueden guardar si el archivo está abierto así que ésta acción finaliza con la operación y define si fue o no exitosa.
        wb.save(Nom_Libro)
        return 1
    except:
        if Iniciado == True:
            return 2
        else:
            return 0

# Convertimos un archivo con pandas
def Convierte_csv_xlsx_Pandas(Path_csv, Destino):
    pd.read_csv(Path_csv).to_excel(Destino, index=False)


print('Módulo Mi_Openpyxl.py cargado correctamente.')
#print(Dev_Listas('D:/Programación/Python/Proyectos/Essen/Anexo4.xlsx','Sheet1', 'H'))

#Convierte_csv_xlsx_Pandas("./contacts.csv", "./Excel_Pandas.xlsx")